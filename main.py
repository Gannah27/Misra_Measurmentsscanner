import openpyxl
from openpyxl.styles import PatternFill
import warnings
import tkinter as tk
from tkinter import ttk
from tkinter import *
from tkinter.filedialog import askopenfile
root = tk.Tk()
filetypes = (
        [('Excel Files', ('*.xlsx'))])
f1=f2=None

def openA_Excel_file():
    global filetypes,f1

    f1 = askopenfile(filetypes=filetypes)
    if(not(f1==None)):
         openN_show.delete(0, END)
         openN_show.insert(0, f1.name)
def openO_Excel_file():
    global filetypes , f2
    f2 = askopenfile( filetypes=filetypes)
    if(not(f2==None)):
        openO_show.delete(0, END)
        openO_show.insert(0, f2.name)

    # read the text file and show its content on the Text

def program():
    global f2,f1
    if(not(f1==None)):
        n_find=f1.name
        print(n_find)
        o_find=f2.name
        print(o_find)
        warnings.simplefilter(action='ignore', category=UserWarning)
        n_obj = openpyxl.load_workbook(n_find)  # All findings sheet
        sheet1_obj = n_obj.active   #All
        o_obj = openpyxl.load_workbook(o_find)  # old findings sheet
        sheet2_obj = o_obj.active  #old
        row_count = sheet1_obj.max_row
        column_count = sheet1_obj.max_column
        old_name=o_obj.sheetnames  #list of sheets name
        count=0
        row_dict= {'Path':[],'Artefact':[],'HIS_CUSTOM_KO':[],	'STAT':[],
                   'HIS_COMF':[],'NPAT':[],'VG':[],'HIS_LEVL':[],'CALI':[],'HIS_CALLS':[],
                   'NOP':[],'GOTO':[],	'RETU':[],'VOCF':[],'CYCL':[],'Cloned Code':[]}

        for i in range(2,row_count+1):
            for j in range(2,column_count+1):
             if sheet1_obj.cell(row=1, column=j).value=='Path':
                 if not(sheet1_obj.cell(row=1, column=j).value):
                     row_dict[sheet1_obj.cell(row=1, column=j).value].append(None)
                 else:
                  row_dict[sheet1_obj.cell(row=1, column=j).value].append(sheet1_obj.cell(row=i, column=17).value.split('/')[2])
             else:
                row_dict[sheet1_obj.cell(row=1, column=j).value].append(sheet1_obj.cell(row=i, column=j).value)
        for r in range(0,row_count-1):
            if(row_dict['Path'][r]=='None'):
                continue
            spec_sheet = o_obj[row_dict['Path'][r]]
            r2 = 0
            col_spec = spec_sheet['B']
            for c in col_spec:
                if not (c.value is None):
                    r2 += 1
            is_found = False
            for k in range(17, r2 + 18):
                scolor=False

                if (spec_sheet.cell(row=k, column=3).value == row_dict['Artefact'][r]):
                    print(row_dict['Artefact'][r])
                    is_found=True
                    for p in range(4 , 19):
                     if (spec_sheet.cell(row=16, column=p).value=='CALD' or spec_sheet.cell(row=16, column=p).value=='CALX'):
                         continue
                     if row_dict[spec_sheet.cell(row=16, column=p).value][r]!=spec_sheet.cell(row=k, column=p).value:
                         sheet1_obj.cell(row=r+2, column= list(row_dict.keys()).index(spec_sheet.cell(row=16, column=p).value)+1).fill = PatternFill(fgColor="5F9EA0", fill_type="solid")
                         scolor=True
                     if scolor:
                         for q in range(1,column_count):
                             if sheet1_obj.cell(row=r+2, column=q).fill != PatternFill(fgColor="5F9EA0", fill_type="solid"):
                                 sheet1_obj.cell(row=r+2, column=q).fill = PatternFill(
                                     fgColor="fff000", fill_type="solid")

                if is_found:
                    break
            if not is_found:
                for q in range(1, column_count):
                    sheet1_obj.cell(row=r+2, column=q).fill = PatternFill(
                    fgColor="ff0000", fill_type="solid")

        n_obj.save(n_find)
        labeldone.place(x=350,y=90)




root.title("Measurements Scanner")
root.configure(bg='light blue')
root.geometry("540x372")
root.resizable(False, False)
tk.Label(root, text="Misra: ", font=('Times', 20), fg="Purple", bg='sky blue').grid(row=0 , column=0)

# Add image file
bg = PhotoImage(file="Brightskies technologies (1).png")

# Show image using label
label1 = Label(root, image=bg)
label1.place(x=0, y=0)
openN_button = ttk.Button(
    root,
    text='Open All Rules File',
    command=openA_Excel_file
)

openN_button.grid(column=0, row=1, sticky='w', padx=10, pady=10)
openN_show = Entry(root)
openN_show.grid(column=1,row=1)
openO_button = ttk.Button(
        root,
    text='Open Reviewed Rules File',
    command=openO_Excel_file
)
openO_show = Entry(root)
openO_show.grid(column=1,row=3)
openO_button.grid(column=0, row=3, sticky='w', padx=10, pady=10)
btn = tk.Button(root, text='Scan files', font=('Times', 12), bg='#a4accc',command=program)
btn.place(x=350,y=50)
labeldone=Label(root,text="Scanning is done")
labelerror=Label(root,text="Note: The scanner only accepts .xlsx extension and workbook type")
label2=Label(root,text="You can change it through file/export/changefile/workbook")
labelerror.place(x=20,y=290)
label2.place(x=20,y=315)
p1=PhotoImage(file='Brightskies technologies (1)-modified.png')
root.iconphoto(False,p1)
root.mainloop()
